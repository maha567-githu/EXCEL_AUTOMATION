from openpyxl import  Workbook,load_workbook
import os
import random
from datetime import datetime
print("CURRENT : ",os.getcwd())
file="EXCEL Automation/Order_Record.xlsx"

# checking file exists or not
if os.path.exists(file):
    ORDER=load_workbook(file)
    sheet=ORDER.active
else:
    ORDER=Workbook()
    sheet=ORDER.active
    sheet.append(["Order_ID","Customer","Pizza","Size","Quantity","Price","Total","Date","Time"])
while True:
    print("1. New Order")  
    print("2. EXIST")  
    Choice=int(input("enter your choice :"))
    if Choice==1:
      Order_id=random.randint(1000,9999)
      Date=datetime.now().strftime("%d-%m-%y")
      Time=datetime.now().strftime("%I:%M %p")
      Customer=input("enter customer name :")
      Fast_Food=input("enter item name:")
      
      while True:
          
          print("1.SMALL")
          print("2.Mediun")
          print("3.EXCEL")
          ch=int(input("enter size :"))
          if ch==1:
              Size="SMALL"
              break
          elif ch==2:
              Size="MEDIUM"
              break
          elif ch==3:
              Size="EXCEL"
              break
          else:
              print("enter valid choice")
      while True:
        try:        
          Quantity=int(input("enter quantity  :"))
          if   Quantity>0:
            break
          else:
             print("enter quantity greater than 0")
        except ValueError:
           print("ENTER NUMBERS ONLY")
      while True:   
        try:  
         Price=int(input("enter price  :"))
         if Price>0:
           break
         else:
           print("enter price greater than 0")
        except ValueError:
           print("NUMBERS ONLY")
      Total=Quantity*Price
      sheet.append([Order_id,Customer,Fast_Food,Size,Quantity,Price,Total,Date,Time])
      ORDER.save(file)
      print("EXCEL FILE GENERATED")
    elif Choice==2:
        print("Come Again")  
        break
    else:
        print("enter a valid choice")